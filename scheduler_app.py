#!/usr/bin/env python3
"""
YRH Internal Medicine Schedule Generator - Web App
Built with Streamlit

To run:
    pip install streamlit pandas openpyxl
    streamlit run scheduler_app.py
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO, StringIO
import random
import csv
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Set, Optional
import zipfile
import re
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import hashlib
import os
from collections.abc import Mapping

# ============================================================================
# PAGE CONFIG
# ============================================================================

st.set_page_config(
    page_title="YRH IM Schedule Generator",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# AUTHENTICATION
# ============================================================================

# Initialize session state for authentication
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.username = None

# Check if debug mode is enabled via query parameter
def is_debug_mode():
    try:
        # Streamlit 1.28+ has query_params
        if hasattr(st, "query_params"):
            return st.query_params.get("debug", "").lower() == "true"
    except:
        pass
    return False

DEBUG = is_debug_mode()

def check_credentials(username, password):
    # Debug: show what's in secrets (will be hidden in production)
    debug_info = []
    
    try:
        # Check Streamlit secrets first
        if hasattr(st, "secrets"):
            debug_info.append(f"st.secrets exists: {type(st.secrets)}")
            if hasattr(st.secrets, "auth"):
                auth = st.secrets.auth
                debug_info.append(f"st.secrets.auth type: {type(auth)}")
                debug_info.append(f"st.secrets.auth content: {auth}")
                
                if isinstance(auth, Mapping):
                    # Format: [auth] admin = "password" (could be dict or AttrDict)
                    debug_info.append(f"Mapping keys: {list(auth.keys())}")
                    if username in auth:
                        debug_info.append(f"Found user '{username}' in mapping")
                        if auth[username] == password:
                            debug_info.append("Password matches!")
                            if DEBUG:
                                st.write("Debug:", debug_info)
                            return True
                        else:
                            debug_info.append("Password mismatch")
                    else:
                        debug_info.append(f"User '{username}' not in mapping")
                elif isinstance(auth, list):
                    # Format: [[auth]] username = "admin" password = "password"
                    debug_info.append(f"List length: {len(auth)}")
                    for i, user in enumerate(auth):
                        debug_info.append(f"List item {i}: {user}")
                        if isinstance(user, dict):
                            if user.get("username") == username:
                                debug_info.append(f"Found user '{username}' in list")
                                if user.get("password") == password:
                                    debug_info.append("Password matches!")
                                    if DEBUG:
                                        st.write("Debug:", debug_info)
                                    return True
                                else:
                                    debug_info.append("Password mismatch")
            else:
                debug_info.append("No 'auth' key in st.secrets")
        else:
            debug_info.append("st.secrets not available")
        
        # Check environment variables
        env_user = os.getenv("SCHEDULER_USERNAME")
        env_pass = os.getenv("SCHEDULER_PASSWORD")
        debug_info.append(f"Env user: {'SET' if env_user else 'NOT SET'}")
        if env_user and env_pass and username == env_user and password == env_pass:
            debug_info.append("Environment variable auth successful")
            if DEBUG:
                st.write("Debug:", debug_info)
            return True
        
        # Hardcoded fallback (for development only)
        if username == "admin" and password == "admin":
            st.warning("Using default credentials - change in production!")
            debug_info.append("Using default admin/admin")
            if DEBUG:
                st.write("Debug:", debug_info)
            return True
        
    except Exception as e:
        st.error(f"Authentication error: {e}")
        debug_info.append(f"Exception: {e}")
        if DEBUG:
            st.write("Debug:", debug_info)
    
    # If we get here, authentication failed
    if DEBUG:
        st.write("Authentication debug:", debug_info)
    return False

# If not authenticated, show login form
if not st.session_state.authenticated:
    st.title("🔒 YRH IM Schedule Generator - Login")
    if DEBUG:
        st.info("🔍 Debug mode enabled. Add ?debug=true to URL to see authentication details.")
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submit = st.form_submit_button("Login")
        if submit:
            if check_credentials(username, password):
                st.session_state.authenticated = True
                st.session_state.username = username
                st.success(f"Welcome, {username}!")
                st.rerun()
            else:
                st.error("Invalid username or password")
    st.stop()

# If authenticated, continue with the app
st.sidebar.markdown(f"👤 **Logged in as:** {st.session_state.username}")

# ============================================================================
# SCHEDULING ENGINE (embedded from im_scheduler_v2.py)
# ============================================================================

@dataclass
class DaySchedule:
    """Schedule for a single day"""
    date: datetime
    im_am: str = ""
    im_pm: str = ""
    icu_am: str = ""
    icu_pm: str = ""
    hospitalist: str = ""
    ecg: str = ""
    cv_clinic: str = ""
    dialysis: str = ""
    stress_test: str = ""
    holidays: dict = field(default_factory=dict)
    
    def is_weekend(self) -> bool:
        return self.date.weekday() in [5, 6]
    
    def is_friday(self) -> bool:
        return self.date.weekday() == 4
    
    def is_holiday(self) -> bool:
        return self.date in self.holidays
    
    def day_name(self) -> str:
        return self.date.strftime("%A")


@dataclass
class DoctorStats:
    """Track assignments for fairness"""
    im_am: int = 0
    im_pm: int = 0
    icu_am: int = 0
    icu_pm: int = 0
    hospitalist: int = 0
    ecg: int = 0
    cv_clinic: int = 0
    dialysis: int = 0
    stress_test: int = 0
    weekends: int = 0
    friday_nights: int = 0
    total_nights: int = 0


class IMScheduler:
    def __init__(self, config: dict):
        self.year = config["year"]
        self.start_date = config.get("start_date") or datetime(self.year, 1, 1)
        self.num_days = config.get("num_days") or (366 if self._is_leap_year(self.year) else 365)
        self.vacations = config.get("vacations", {})
        self.holidays = config.get("holidays", {})
        self.use_historical = config.get("use_historical", False)
        self.historical_counts = config.get("historical_counts", {})
        
        # Blair-type rules: list of {"doctor": name, "interval": weeks, "first_friday": datetime}
        self.blair_rules = config.get("blair_rules", [])
        
        # Week start days (0=Monday ... 6=Sunday)
        self.im_icu_week_start = config.get("im_icu_week_start", 4)  # Default: Friday
        self.dialysis_week_start = config.get("dialysis_week_start", 0)  # Default: Monday
        
        # Set up doctors from config
        self._setup_doctors_from_config(
            config.get("doctors", {}), 
            config.get("blair_doctor", "MacDonald")  # Legacy support
        )
        
        self.schedule: List[DaySchedule] = []
        self.stats: Dict[str, DoctorStats] = {doc: DoctorStats() for doc in self.all_doctors}
        
        # Initialize with historical counts if provided
        if self.use_historical and self.historical_counts:
            for doc, counts in self.historical_counts.items():
                if doc in self.stats:
                    self.stats[doc].im_am = counts.get("IM_AM", 0)
                    self.stats[doc].im_pm = counts.get("IM_PM", 0)
                    self.stats[doc].icu_am = counts.get("ICU_AM", 0)
                    self.stats[doc].icu_pm = counts.get("ICU_PM", 0)
                    self.stats[doc].hospitalist = counts.get("Hospitalist", 0)
                    self.stats[doc].ecg = counts.get("ECG", 0)
                    self.stats[doc].cv_clinic = counts.get("CVClinic", 0)
                    self.stats[doc].dialysis = counts.get("Dialysis", 0)
                    self.stats[doc].stress_test = counts.get("StressTest", 0)
                    self.stats[doc].weekends = counts.get("Weekends", 0)
                    self.stats[doc].total_nights = counts.get("Nights", 0)
        
        # Initialize schedule
        for i in range(self.num_days):
            day = DaySchedule(date=self.start_date + timedelta(days=i), holidays=self.holidays)
            self.schedule.append(day)
    
    def _setup_doctors_from_config(self, doctors_config: dict, blair_doctor: str):
        """Set up doctor lists and service coverage from config"""
        self.all_doctors = list(doctors_config.keys())
        self.blair_doctor = blair_doctor
        
        self.service_coverage = {
            "ICU": [],
            "IM": [],
            "Hospitalist": [],
            "StressTest": [],
            "CVClinic": [],
            "ECG": [],
            "Dialysis": [],
        }
        
        for doctor, services in doctors_config.items():
            for service in services:
                service_upper = service.upper()
                if service_upper == "ICU":
                    self.service_coverage["ICU"].append(doctor)
                elif service_upper == "IM":
                    self.service_coverage["IM"].append(doctor)
                elif service_upper == "HOSPITALIST":
                    self.service_coverage["Hospitalist"].append(doctor)
                elif service_upper in ["STRESSTEST", "STRESS"]:
                    self.service_coverage["StressTest"].append(doctor)
                elif service_upper in ["CVCLINIC", "CV"]:
                    self.service_coverage["CVClinic"].append(doctor)
                elif service_upper == "ECG":
                    self.service_coverage["ECG"].append(doctor)
                elif service_upper == "DIALYSIS":
                    self.service_coverage["Dialysis"].append(doctor)
        
        icu_set = set(self.service_coverage["ICU"])
        im_set = set(self.service_coverage["IM"])
        hosp_set = set(self.service_coverage["Hospitalist"])
        
        self.dual_coverage = list(icu_set & im_set)
        self.icu_only = list(icu_set - im_set)
        self.im_only = list(im_set - icu_set)
        self.icu_hospitalist = list(icu_set & hosp_set)
    
    def _is_leap_year(self, year: int) -> bool:
        return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
    
    def _calculate_blair_start(self) -> datetime:
        base = datetime(2026, 1, 2)
        year_start = datetime(self.year, 1, 1)
        first_friday = year_start
        while first_friday.weekday() != 4:
            first_friday += timedelta(days=1)
        weeks_diff = (first_friday - base).days // 7
        offset = weeks_diff % 5
        if offset == 0:
            return first_friday
        else:
            return first_friday + timedelta(weeks=(5 - offset))
    
    def is_on_vacation(self, doctor: str, date: datetime) -> bool:
        if doctor not in self.vacations:
            return False
        return date in self.vacations[doctor]
    
    def get_available_doctors(self, service: str, date: datetime, exclude: Set[str] = None) -> List[str]:
        exclude = exclude or set()
        available = []
        for doc in self.service_coverage.get(service, []):
            if doc not in exclude and not self.is_on_vacation(doc, date):
                available.append(doc)
        return available
    
    def get_week_start_friday(self, date: datetime) -> datetime:
        """Get the start of the IM/ICU week based on configured week start day"""
        days_since_start = (date.weekday() - self.im_icu_week_start) % 7
        return date - timedelta(days=days_since_start)
    
    def get_monday_week_start(self, date: datetime) -> datetime:
        """Get the start of the Dialysis week based on configured week start day"""
        days_since_start = (date.weekday() - self.dialysis_week_start) % 7
        return date - timedelta(days=days_since_start)
    
    def is_blair_icu_week(self, date: datetime) -> bool:
        """Check if any Blair-rule doctor is on ICU this week, return the doctor or None"""
        return self.get_blair_doctor_for_week(date) is not None
    
    def get_blair_doctor_for_week(self, date: datetime) -> Optional[str]:
        """Get which Blair-rule doctor (if any) is on ICU for this week"""
        week_start = self.get_week_start_friday(date)
        
        for rule in self.blair_rules:
            doctor = rule["doctor"]
            interval = rule["interval"]
            first_friday = rule["first_friday"]
            
            if isinstance(first_friday, str):
                first_friday = datetime.strptime(first_friday, "%Y-%m-%d")
            
            weeks_since_first = (week_start - first_friday).days // 7
            if weeks_since_first >= 0 and weeks_since_first % interval == 0:
                return doctor
        
        return None
    
    def select_doctor_by_fairness(self, candidates: List[str], stat_attr: str, 
                                   date: datetime, is_weekend: bool = False,
                                   is_night: bool = False, is_friday_night: bool = False) -> str:
        if not candidates:
            return ""
        
        scores = []
        for doc in candidates:
            stat = self.stats[doc]
            primary_score = getattr(stat, stat_attr, 0)
            weekend_penalty = stat.weekends * 0.5 if is_weekend else 0
            night_penalty = stat.total_nights * 0.3 if is_night else 0
            friday_penalty = stat.friday_nights * 0.4 if is_friday_night else 0
            total_score = primary_score + weekend_penalty + night_penalty + friday_penalty
            scores.append((doc, total_score, primary_score))
        
        scores.sort(key=lambda x: (x[1], x[2]))
        top_score = scores[0][1]
        similar = [s for s in scores if s[1] <= top_score + 2]
        
        if len(similar) > 1:
            return random.choice(similar)[0]
        return scores[0][0]
    
    def assign_weekly_services(self):
        current_im_am = None
        current_icu_am = None
        current_hospitalist = None
        current_dialysis = None
        current_week_start = None
        current_dialysis_week_start = None
        
        # Get all Blair-rule doctors
        blair_doctors = set(rule["doctor"] for rule in self.blair_rules)
        
        for day in self.schedule:
            date = day.date
            week_start = self.get_week_start_friday(date)
            dialysis_week_start = self.get_monday_week_start(date)
            
            if week_start != current_week_start:
                current_week_start = week_start
                
                # Check if any Blair-rule doctor is on ICU this week
                blair_doc_this_week = self.get_blair_doctor_for_week(date)
                
                if blair_doc_this_week:
                    current_icu_am = blair_doc_this_week
                else:
                    # Exclude all Blair-rule doctors from regular rotation
                    icu_candidates = [d for d in self.service_coverage["ICU"] 
                                     if d not in blair_doctors and not self.is_on_vacation(d, date)]
                    if icu_candidates:
                        current_icu_am = self.select_doctor_by_fairness(
                            icu_candidates, "icu_am", date, is_weekend=day.is_weekend())
                
                im_candidates = self.get_available_doctors("IM", date)
                im_candidates = [d for d in im_candidates if d != current_icu_am]
                
                if current_icu_am in self.icu_only:
                    im_preferred = [d for d in im_candidates if d in self.dual_coverage]
                    if not im_preferred:
                        im_preferred = [d for d in im_candidates if d in self.im_only]
                    if not im_preferred:
                        im_preferred = im_candidates
                else:
                    im_preferred = [d for d in im_candidates if d in self.im_only]
                    if not im_preferred:
                        im_preferred = im_candidates
                
                if im_preferred:
                    current_im_am = self.select_doctor_by_fairness(
                        im_preferred, "im_am", date, is_weekend=day.is_weekend())
                elif im_candidates:
                    current_im_am = self.select_doctor_by_fairness(
                        im_candidates, "im_am", date, is_weekend=day.is_weekend())
                
                if current_icu_am in self.icu_hospitalist:
                    current_hospitalist = current_icu_am
                else:
                    hosp_candidates = self.get_available_doctors("Hospitalist", date, 
                                                                  exclude={current_im_am, current_icu_am})
                    if hosp_candidates:
                        current_hospitalist = self.select_doctor_by_fairness(
                            hosp_candidates, "hospitalist", date, is_weekend=day.is_weekend())
            
            if dialysis_week_start != current_dialysis_week_start:
                current_dialysis_week_start = dialysis_week_start
                dialysis_candidates = self.get_available_doctors("Dialysis", date)
                if dialysis_candidates:
                    current_dialysis = self.select_doctor_by_fairness(dialysis_candidates, "dialysis", date)
            
            if current_icu_am and not self.is_on_vacation(current_icu_am, date):
                day.icu_am = current_icu_am
                self.stats[current_icu_am].icu_am += 1
                if day.is_weekend():
                    self.stats[current_icu_am].weekends += 1
            
            if current_im_am and not self.is_on_vacation(current_im_am, date):
                day.im_am = current_im_am
                self.stats[current_im_am].im_am += 1
                day.ecg = current_im_am
                self.stats[current_im_am].ecg += 1
                if day.is_weekend():
                    self.stats[current_im_am].weekends += 1
            
            if current_hospitalist and not self.is_on_vacation(current_hospitalist, date):
                day.hospitalist = current_hospitalist
                self.stats[current_hospitalist].hospitalist += 1
            
            if current_dialysis and not self.is_on_vacation(current_dialysis, date):
                day.dialysis = current_dialysis
                self.stats[current_dialysis].dialysis += 1
    
    def assign_night_calls(self):
        consecutive_nights = {doc: 0 for doc in self.all_doctors}
        last_night_date = {doc: None for doc in self.all_doctors}
        
        for day in self.schedule:
            date = day.date
            day_name = day.day_name()
            is_friday_night = day_name == "Friday"
            icu_day_doc = day.icu_am
            im_day_doc = day.im_am
            
            for doc in self.all_doctors:
                if last_night_date[doc]:
                    gap = (date - last_night_date[doc]).days
                    if gap > 1:
                        consecutive_nights[doc] = 0
            
            def can_do_night(doc: str) -> bool:
                if not doc:
                    return False
                if self.is_on_vacation(doc, date):
                    return False
                if last_night_date[doc] and (date - last_night_date[doc]).days == 1:
                    if consecutive_nights[doc] >= 3:
                        return False
                return True
            
            def record_night(doc: str):
                if not doc:
                    return
                if last_night_date[doc] == date:
                    return
                if last_night_date[doc] and (date - last_night_date[doc]).days == 1:
                    consecutive_nights[doc] += 1
                else:
                    consecutive_nights[doc] = 1
                last_night_date[doc] = date
            
            if day_name in ["Friday", "Sunday", "Tuesday", "Thursday"]:
                primary_service = "ICU"
            else:
                primary_service = "IM"
            
            # Check if a Blair-rule doctor is on ICU this week
            blair_doc_this_week = self.get_blair_doctor_for_week(date)
            is_blair_night = False
            if blair_doc_this_week:
                if day_name in ["Friday", "Saturday", "Sunday", "Tuesday", "Thursday"]:
                    is_blair_night = True
            
            icu_night_doc = None
            
            if is_blair_night and blair_doc_this_week and can_do_night(blair_doc_this_week):
                icu_night_doc = blair_doc_this_week
            elif primary_service == "ICU" and can_do_night(icu_day_doc):
                icu_night_doc = icu_day_doc
            elif primary_service == "IM" and im_day_doc in self.dual_coverage and can_do_night(im_day_doc):
                icu_night_doc = im_day_doc
            else:
                icu_candidates = [d for d in self.service_coverage["ICU"] 
                                 if can_do_night(d) and not self.is_on_vacation(d, date)]
                dual_icu = [d for d in icu_candidates if d in self.dual_coverage]
                if dual_icu:
                    icu_night_doc = self.select_doctor_by_fairness(
                        dual_icu, "icu_pm", date, is_night=True, is_friday_night=is_friday_night)
                elif icu_candidates:
                    icu_night_doc = self.select_doctor_by_fairness(
                        icu_candidates, "icu_pm", date, is_night=True, is_friday_night=is_friday_night)
            
            if icu_night_doc:
                day.icu_pm = icu_night_doc
                self.stats[icu_night_doc].icu_pm += 1
                self.stats[icu_night_doc].total_nights += 1
                if is_friday_night:
                    self.stats[icu_night_doc].friday_nights += 1
                record_night(icu_night_doc)
            
            im_night_doc = None
            
            if icu_night_doc and icu_night_doc in self.dual_coverage:
                im_night_doc = icu_night_doc
            elif primary_service == "IM" and can_do_night(im_day_doc):
                im_night_doc = im_day_doc
            elif primary_service == "ICU" and icu_day_doc in self.dual_coverage and can_do_night(icu_day_doc):
                im_night_doc = icu_day_doc
            else:
                im_candidates = [d for d in self.service_coverage["IM"]
                               if can_do_night(d) and not self.is_on_vacation(d, date)
                               and d != icu_night_doc]
                if im_candidates:
                    im_night_doc = self.select_doctor_by_fairness(
                        im_candidates, "im_pm", date, is_night=True, is_friday_night=is_friday_night)
            
            if im_night_doc:
                day.im_pm = im_night_doc
                self.stats[im_night_doc].im_pm += 1
                if im_night_doc != icu_night_doc:
                    self.stats[im_night_doc].total_nights += 1
                    if is_friday_night:
                        self.stats[im_night_doc].friday_nights += 1
                record_night(im_night_doc)
    
    def assign_cv_clinic(self):
        for day in self.schedule:
            if day.is_friday() and not day.is_holiday():
                exclude = {day.im_am, day.icu_am}
                if day.hospitalist != day.icu_am:
                    exclude.add(day.hospitalist)
                
                candidates = self.get_available_doctors("CVClinic", day.date, exclude=exclude)
                if not candidates:
                    candidates = self.get_available_doctors("CVClinic", day.date)
                
                if candidates:
                    cv_doc = self.select_doctor_by_fairness(candidates, "cv_clinic", day.date)
                    day.cv_clinic = cv_doc
                    self.stats[cv_doc].cv_clinic += 1
    
    def assign_stress_tests(self):
        for day in self.schedule:
            if day.day_name() in ["Monday", "Tuesday", "Wednesday", "Thursday"]:
                if day.is_holiday():
                    day.stress_test = "HOLIDAY"
                else:
                    exclude = {day.im_am, day.icu_am}
                    if day.hospitalist != day.icu_am:
                        exclude.add(day.hospitalist)
                    
                    candidates = self.get_available_doctors("StressTest", day.date, exclude=exclude)
                    if not candidates:
                        candidates = self.get_available_doctors("StressTest", day.date)
                    
                    if candidates:
                        st_doc = self.select_doctor_by_fairness(candidates, "stress_test", day.date)
                        day.stress_test = st_doc
                        self.stats[st_doc].stress_test += 1
            elif day.is_friday():
                day.stress_test = "CV Clinic"
            elif day.is_weekend():
                day.stress_test = "WEEKEND"
    
    def generate_schedule(self) -> List[DaySchedule]:
        self.assign_weekly_services()
        self.assign_night_calls()
        self.assign_cv_clinic()
        self.assign_stress_tests()
        return self.schedule
    
    def get_fairness_report(self) -> str:
        report = []
        report.append("=" * 100)
        report.append(f"FAIRNESS STATISTICS REPORT - {self.year}")
        report.append("=" * 100 + "\n")
        
        headers = ["Doctor", "IM AM", "IM PM", "ICU AM", "ICU PM", "Hosp", "ECG", 
                   "CV", "Dial", "Stress", "Wknd", "Nights", "Fri Ngt"]
        
        report.append(f"{headers[0]:<15}" + "".join(f"{h:>8}" for h in headers[1:]))
        report.append("-" * 115)
        
        for doc in sorted(self.all_doctors):
            stat = self.stats[doc]
            report.append(
                f"{doc:<15}"
                f"{stat.im_am:>8}"
                f"{stat.im_pm:>8}"
                f"{stat.icu_am:>8}"
                f"{stat.icu_pm:>8}"
                f"{stat.hospitalist:>8}"
                f"{stat.ecg:>8}"
                f"{stat.cv_clinic:>8}"
                f"{stat.dialysis:>8}"
                f"{stat.stress_test:>8}"
                f"{stat.weekends:>8}"
                f"{stat.total_nights:>8}"
                f"{stat.friday_nights:>8}"
            )
        
        return "\n".join(report)
    
    def check_conflicts(self) -> List[str]:
        conflicts = []
        consecutive_night = {doc: 0 for doc in self.all_doctors}
        last_night_date = {doc: None for doc in self.all_doctors}
        
        for day in self.schedule:
            date = day.date
            date_str = date.strftime("%a %m/%d")
            
            for service, doc in [
                ("IM AM", day.im_am),
                ("IM PM", day.im_pm),
                ("ICU AM", day.icu_am),
                ("ICU PM", day.icu_pm),
                ("Hospitalist", day.hospitalist),
                ("Dialysis", day.dialysis),
            ]:
                if doc and doc in self.all_doctors and self.is_on_vacation(doc, date):
                    conflicts.append(f"CONFLICT [{date_str}]: {doc} scheduled for {service} but is on VACATION")
            
            if not day.icu_am:
                conflicts.append(f"CONFLICT [{date_str}]: No ICU AM coverage")
            if not day.icu_pm:
                conflicts.append(f"CONFLICT [{date_str}]: No ICU PM coverage")
            if not day.im_am:
                conflicts.append(f"CONFLICT [{date_str}]: No IM AM coverage")
            if not day.im_pm:
                conflicts.append(f"CONFLICT [{date_str}]: No IM PM coverage")
            if not day.hospitalist:
                conflicts.append(f"CONFLICT [{date_str}]: No Hospitalist coverage")
            if not day.dialysis:
                conflicts.append(f"CONFLICT [{date_str}]: No Dialysis coverage")
            
            night_docs_today = set()
            if day.icu_pm and day.icu_pm in self.all_doctors:
                night_docs_today.add(day.icu_pm)
            if day.im_pm and day.im_pm in self.all_doctors:
                night_docs_today.add(day.im_pm)
            
            for doc in night_docs_today:
                if last_night_date[doc] and (date - last_night_date[doc]).days == 1:
                    consecutive_night[doc] += 1
                    if consecutive_night[doc] > 3:
                        conflicts.append(f"WARNING [{date_str}]: {doc} on {consecutive_night[doc]} consecutive nights (max 3)")
                else:
                    consecutive_night[doc] = 1
                last_night_date[doc] = date
            
            for doc in self.all_doctors:
                if doc not in night_docs_today:
                    if last_night_date[doc] and (date - last_night_date[doc]).days > 1:
                        consecutive_night[doc] = 0
        
        return conflicts
    
    def generate_conflict_report(self) -> str:
        conflicts = self.check_conflicts()
        
        report = []
        report.append("=" * 80)
        report.append("CONFLICT & VALIDATION REPORT")
        report.append("=" * 80)
        
        if not conflicts:
            report.append("\n✓ No conflicts or warnings found! Schedule is valid.")
        else:
            conflict_list = [c for c in conflicts if c.startswith("CONFLICT")]
            warning_list = [c for c in conflicts if c.startswith("WARNING")]
            
            report.append(f"\nFound {len(conflict_list)} conflicts and {len(warning_list)} warnings:\n")
            
            if conflict_list:
                report.append("CONFLICTS (must fix):")
                report.append("-" * 40)
                for c in conflict_list[:50]:
                    report.append(f"  {c}")
                if len(conflict_list) > 50:
                    report.append(f"  ... and {len(conflict_list) - 50} more conflicts")
            
            if warning_list:
                report.append("\nWARNINGS (review recommended):")
                report.append("-" * 40)
                for w in warning_list[:50]:
                    report.append(f"  {w}")
                if len(warning_list) > 50:
                    report.append(f"  ... and {len(warning_list) - 50} more warnings")
        
        report.append("\n" + "=" * 80)
        return "\n".join(report)
    
    def to_dataframe(self) -> pd.DataFrame:
        data = []
        for day in self.schedule:
            data.append({
                "Day": day.date.strftime("%A %B %d, %Y"),
                "IM On Call 8am-5pm": day.im_am,
                "IM On Call 5pm-8am": day.im_pm,
                "ICU Days": day.icu_am,
                "ICU Nights": day.icu_pm,
                "YRH Hospitalist": day.hospitalist,
                "ECGs": day.ecg,
                "CV Clinic": day.cv_clinic if day.cv_clinic else "-",
                "Dialysis": day.dialysis,
                "Stress Tests": day.stress_test,
            })
        return pd.DataFrame(data)


# ============================================================================
# DEFAULT DATA
# ============================================================================

DEFAULT_DOCTORS = {
    "Eissa": ["ICU", "IM", "Hospitalist", "ECG", "StressTest", "CVClinic"],
    "Janzen": ["ICU", "IM", "Hospitalist", "ECG", "StressTest", "CVClinic"],
    "Moses": ["ICU", "IM", "Hospitalist", "ECG", "StressTest", "CVClinic", "Dialysis"],
    "Chahal": ["ICU", "IM", "Hospitalist", "ECG", "StressTest", "CVClinic", "Dialysis"],
    "KhierElseed": ["IM", "Hospitalist", "ECG", "Dialysis"],
    "Rao-Ullur": ["IM", "Hospitalist", "ECG", "Dialysis"],
    "Elsheaita": ["IM", "Hospitalist", "ECG"],
    "MacDonald": ["ICU"],
    "McLaren-Blades": ["ICU"],
    "Saeid": ["StressTest", "CVClinic"],
}

ALL_SERVICES = ["ICU", "IM", "Hospitalist", "ECG", "StressTest", "CVClinic", "Dialysis"]

def calculate_easter(year: int) -> datetime:
    """Calculate Easter Sunday using the Anonymous Gregorian algorithm"""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return datetime(year, month, day)


def get_nth_weekday(year: int, month: int, weekday: int, n: int) -> datetime:
    """Get the nth occurrence of a weekday in a month (n=1 for first, n=-1 for last)"""
    if n > 0:
        # First day of month
        first_day = datetime(year, month, 1)
        # Days until first occurrence of weekday
        days_until = (weekday - first_day.weekday()) % 7
        first_occurrence = first_day + timedelta(days=days_until)
        return first_occurrence + timedelta(weeks=n-1)
    else:
        # Last day of month
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        # Days back to last occurrence of weekday
        days_back = (last_day.weekday() - weekday) % 7
        return last_day - timedelta(days=days_back)


def generate_ns_holidays(year: int) -> dict:
    """Generate Nova Scotia statutory holidays for a given year"""
    holidays = {}
    
    # Fixed date holidays
    holidays[f"{year}-01-01"] = "New Year's Day"
    holidays[f"{year}-07-01"] = "Canada Day"
    holidays[f"{year}-11-11"] = "Remembrance Day"
    holidays[f"{year}-12-25"] = "Christmas Day"
    holidays[f"{year}-12-26"] = "Boxing Day"
    
    # Nova Scotia Heritage Day - 3rd Monday of February
    heritage_day = get_nth_weekday(year, 2, 0, 3)  # 0 = Monday
    holidays[heritage_day.strftime("%Y-%m-%d")] = "Heritage Day"
    
    # Easter-based holidays
    easter = calculate_easter(year)
    good_friday = easter - timedelta(days=2)
    easter_monday = easter + timedelta(days=1)
    holidays[good_friday.strftime("%Y-%m-%d")] = "Good Friday"
    holidays[easter_monday.strftime("%Y-%m-%d")] = "Easter Monday"
    
    # Victoria Day - Monday before May 25
    may_25 = datetime(year, 5, 25)
    days_since_monday = (may_25.weekday()) % 7
    if days_since_monday == 0:
        victoria_day = may_25
    else:
        victoria_day = may_25 - timedelta(days=days_since_monday)
    holidays[victoria_day.strftime("%Y-%m-%d")] = "Victoria Day"
    
    # Civic Holiday (Natal Day in NS) - 1st Monday of August
    civic_holiday = get_nth_weekday(year, 8, 0, 1)
    holidays[civic_holiday.strftime("%Y-%m-%d")] = "Civic Holiday"
    
    # Labour Day - 1st Monday of September
    labour_day = get_nth_weekday(year, 9, 0, 1)
    holidays[labour_day.strftime("%Y-%m-%d")] = "Labour Day"
    
    # National Day for Truth and Reconciliation - September 30 (since 2021)
    holidays[f"{year}-09-30"] = "Truth and Reconciliation Day"
    
    # Thanksgiving - 2nd Monday of October
    thanksgiving = get_nth_weekday(year, 10, 0, 2)
    holidays[thanksgiving.strftime("%Y-%m-%d")] = "Thanksgiving"
    
    return holidays


# Default holidays - will be generated based on selected year
NS_HOLIDAYS_2026 = generate_ns_holidays(2026)
NS_HOLIDAYS_2027 = generate_ns_holidays(2027)


# ============================================================================
# SESSION STATE INITIALIZATION
# ============================================================================

if "doctors" not in st.session_state:
    st.session_state.doctors = DEFAULT_DOCTORS.copy()

if "vacations" not in st.session_state:
    st.session_state.vacations = {}

if "holidays" not in st.session_state:
    st.session_state.holidays = NS_HOLIDAYS_2027.copy()

if "generated_schedules" not in st.session_state:
    st.session_state.generated_schedules = None

# Blair-type rules: list of doctors with special rotation patterns
# Format: {"doctor": name, "interval": weeks, "first_friday": date}
if "blair_rules" not in st.session_state:
    st.session_state.blair_rules = [
        {"doctor": "MacDonald", "interval": 5, "first_friday": "2027-01-08"}
    ]

# Week start days (0=Monday, 1=Tuesday, ..., 4=Friday, 5=Saturday, 6=Sunday)
if "week_starts" not in st.session_state:
    st.session_state.week_starts = {
        "IM_ICU": 4,      # Friday (current default)
        "Dialysis": 0,    # Monday (current default)
    }


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def parse_vacation_csv(uploaded_file) -> dict:
    """Parse uploaded vacation CSV/Excel file"""
    vacations = defaultdict(list)
    
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
        
        # Normalize column names
        df.columns = [c.strip().lower() for c in df.columns]
        
        # Find doctor column
        doc_col = None
        for col in ['doctor', 'name', 'physician']:
            if col in df.columns:
                doc_col = col
                break
        
        if doc_col is None:
            doc_col = df.columns[0]
        
        # Find date columns
        start_col = None
        end_col = None
        for col in df.columns:
            if 'start' in col:
                start_col = col
            elif 'end' in col:
                end_col = col
        
        if start_col is None:
            start_col = df.columns[1] if len(df.columns) > 1 else None
        if end_col is None:
            end_col = df.columns[2] if len(df.columns) > 2 else None
        
        for _, row in df.iterrows():
            doctor = str(row[doc_col]).strip()
            if not doctor or doctor == 'nan':
                continue
            
            try:
                start = pd.to_datetime(row[start_col])
                end = pd.to_datetime(row[end_col]) if end_col else start
                
                current = start
                while current <= end:
                    vacations[doctor].append(current.to_pydatetime())
                    current += timedelta(days=1)
            except:
                continue
        
        return dict(vacations)
    except Exception as e:
        st.error(f"Error parsing file: {e}")
        return {}


def parse_historical_report(uploaded_file) -> dict:
    """Parse a fairness report to extract historical counts"""
    historical = {}
    
    try:
        content = uploaded_file.read().decode('utf-8')
        lines = content.split('\n')
        
        # Find the data rows (skip headers)
        in_data = False
        for line in lines:
            line = line.strip()
            
            # Skip empty lines and separators
            if not line or line.startswith('=') or line.startswith('-'):
                continue
            
            # Skip header lines
            if 'Doctor' in line and ('IM AM' in line or 'IM_AM' in line):
                in_data = True
                continue
            
            if 'FAIRNESS' in line or 'REPORT' in line:
                continue
            
            if in_data:
                # Try to parse as data row
                parts = line.split()
                if len(parts) >= 10:
                    try:
                        doctor = parts[0]
                        # Check if first part is a name (not a number)
                        if not doctor[0].isdigit():
                            historical[doctor] = {
                                "IM_AM": int(parts[1]) if parts[1].isdigit() else 0,
                                "IM_PM": int(parts[2]) if parts[2].isdigit() else 0,
                                "ICU_AM": int(parts[3]) if parts[3].isdigit() else 0,
                                "ICU_PM": int(parts[4]) if parts[4].isdigit() else 0,
                                "Hospitalist": int(parts[5]) if parts[5].isdigit() else 0,
                                "ECG": int(parts[6]) if parts[6].isdigit() else 0,
                                "CVClinic": int(parts[7]) if parts[7].isdigit() else 0,
                                "Dialysis": int(parts[8]) if parts[8].isdigit() else 0,
                                "StressTest": int(parts[9]) if parts[9].isdigit() else 0,
                                "Weekends": int(parts[10]) if len(parts) > 10 and parts[10].isdigit() else 0,
                                "Nights": int(parts[11]) if len(parts) > 11 and parts[11].isdigit() else 0,
                            }
                    except (ValueError, IndexError):
                        continue
        
        return historical if historical else None
        
    except Exception as e:
        st.error(f"Error parsing historical report: {e}")
        return None


def doctors_to_dataframe(doctors: dict) -> pd.DataFrame:
    """Convert doctors dict to DataFrame for editing"""
    rows = []
    for name, services in doctors.items():
        row = {"Doctor": name}
        for svc in ALL_SERVICES:
            row[svc] = svc in services
        rows.append(row)
    return pd.DataFrame(rows)


def dataframe_to_doctors(df: pd.DataFrame) -> dict:
    """Convert DataFrame back to doctors dict"""
    doctors = {}
    for _, row in df.iterrows():
        name = row["Doctor"]
        if not name or str(name).strip() == "":
            continue
        services = [svc for svc in ALL_SERVICES if row.get(svc, False)]
        doctors[name] = services
    return doctors


def _google_sheet_export_csv_url(sheet_url: str, gid: str | None = None) -> str:
    """Build a CSV export URL from a Google Sheets sharing URL."""
    # Accept either full /edit URL or already-export URL.
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
    if not m:
        raise ValueError("Could not parse spreadsheet id from URL")
    sid = m.group(1)
    params = "format=csv"
    if gid:
        params += f"&gid={gid}"
    return f"https://docs.google.com/spreadsheets/d/{sid}/export?{params}"


def load_historical_counts_from_schedule_df(df: pd.DataFrame, doctors: list[str], cutoff_date: datetime | None = None) -> dict:
    """Compute historical service counts per doctor from an existing schedule sheet DF.

    If cutoff_date is provided, only rows with Day < cutoff_date are counted.

    Expects columns like:
      - IM On Call 8am-5pm
      - IM On Call 5pm-8am
      - ICU Days
      - ICU Nights
      - YRH Hospitalist
      - ECGs
      - CV Clinic
      - Dialysis
      - Stress Tests

    Returns dict matching parse_historical_report format.
    """
    # column name normalization (loose matching)
    col_map = {
        'IM_AM': ['IM On Call 8am-5pm', 'IM On Call 8am-5pm ', 'IM On Call 8am-5pm\n'],
        'IM_PM': ['IM On Call 5pm-8am'],
        'ICU_AM': ['ICU Days'],
        'ICU_PM': ['ICU Nights'],
        'Hospitalist': ['YRH Hospitalist'],
        'ECG': ['ECGs', 'ECG'],
        'CVClinic': ['CV Clinic', 'CVClinic'],
        'Dialysis': ['Dialysis'],
        'StressTest': ['Stress Tests', 'StressTest'],
    }

    def pick_col(possibles: list[str]) -> str | None:
        for c in possibles:
            if c in df.columns:
                return c
        # fallback: case-insensitive contains
        for c in df.columns:
            for p in possibles:
                if p.lower() in str(c).lower():
                    return c
        return None

    cols = {k: pick_col(v) for k, v in col_map.items()}

    # Init counts
    out = {doc: {
        'IM_AM': 0, 'IM_PM': 0, 'ICU_AM': 0, 'ICU_PM': 0,
        'Hospitalist': 0, 'ECG': 0, 'CVClinic': 0, 'Dialysis': 0, 'StressTest': 0,
        'Weekends': 0, 'Nights': 0,
    } for doc in doctors}

    def inc(doc: str, key: str, amount: int = 1):
        if doc in out:
            out[doc][key] = out[doc].get(key, 0) + amount

    # helper: normalize a cell to a doctor token
    def cell_doctor(v) -> str | None:
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.lower() in ['nan', '-', 'tbd', 'none']:
            return None
        # many cells have combos like "Eissa/Howatt" or "Rao-Ullur-Khierelseed"
        # we count each listed doctor if present in doctor list.
        return s

    # Weekend detection if a Day column exists
    day_col = None
    for c in df.columns:
        if str(c).lower().startswith('day'):
            day_col = c
            break

    for _, row in df.iterrows():
        is_weekend = False
        row_dt = None
        if day_col:
            try:
                row_dt = pd.to_datetime(row[day_col])
                # cutoff filter
                if cutoff_date is not None and pd.notna(row_dt) and row_dt.to_pydatetime() >= cutoff_date:
                    continue
                is_weekend = row_dt.weekday() in [5, 6]
            except:
                row_dt = None

        # day roles
        for key, col in cols.items():
            if not col:
                continue
            raw = cell_doctor(row.get(col))
            if not raw:
                continue

            # split on common separators
            parts = re.split(r"[\/,&]+|\s+and\s+|\s*-\s*", raw)
            parts = [p.strip() for p in parts if p.strip()]

            if key in ['ICU_PM', 'IM_PM']:
                # count nights
                for p in parts:
                    if p in out:
                        inc(p, key)
                        inc(p, 'Nights')
                        if is_weekend:
                            inc(p, 'Weekends')
            else:
                for p in parts:
                    if p in out:
                        inc(p, key)
                        if is_weekend:
                            inc(p, 'Weekends')

    # remove empty doctors (not in DF at all)
    return out


def fetch_google_sheet_schedule_counts(sheet_url: str, gid: str | None, doctors: list[str], cutoff_date: datetime | None = None) -> dict:
    """Fetch schedule from Google Sheet CSV export and return historical counts.

    If cutoff_date is provided, only counts rows before that date.
    """
    csv_url = _google_sheet_export_csv_url(sheet_url, gid)
    r = requests.get(csv_url, timeout=20)
    r.raise_for_status()
    content = r.content.decode('utf-8', errors='replace')
    df = pd.read_csv(StringIO(content))
    return load_historical_counts_from_schedule_df(df, doctors, cutoff_date=cutoff_date)


def export_schedule_to_excel(scheduler, filename: str = None) -> BytesIO:
    """
    Export the schedule to a formatted Excel workbook with:
      - Sheet 1: Schedule grid with doctor-specific colors and today's row highlighted.
      - Sheet 2: Counts & fairness metrics.
      - Sheet 3: Doctor-color legend.
    Returns BytesIO object.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from datetime import datetime as dt

    # Create workbook
    wb = Workbook()
    ws_schedule = wb.active
    ws_schedule.title = "Schedule"
    
    # Get dataframe
    df = scheduler.to_dataframe()
    
    # Map doctor names to consistent colors
    doctor_colors = {}
    for doc in sorted(scheduler.all_doctors):
        # Generate a consistent color from doctor name hash
        hash_int = int(hashlib.md5(doc.encode()).hexdigest()[:8], 16)
        hue = hash_int % 360
        # Use HSL to RGB conversion (simplified: use a palette of distinct colors)
        # Instead, use a fixed palette of 20 colors, cycle if more doctors
        palette = [
            "FFB6C1", "FFD700", "98FB98", "87CEEB", "DDA0DD", "F0E68C", "E6E6FA", 
            "FFA07A", "20B2AA", "FF6347", "7B68EE", "00FA9A", "FFDAB9", "FF69B4", 
            "BA55D3", "40E0D0", "FF8C00", "778899", "32CD32", "6A5ACD"
        ]
        idx = hash_int % len(palette)
        doctor_colors[doc] = palette[idx]
    
    # Write header
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws_schedule.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write data rows
    today = dt.now().date()
    today_row = None
    
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        # Check if this row is today's date
        day_str = row[0]  # "Day" column
        # Parse date from string like "Monday January 01, 2025"
        try:
            # Try multiple date formats
            parsed_date = None
            for fmt in ("%A %B %d, %Y", "%A %b %d, %Y", "%A %d %B %Y"):
                try:
                    parsed_date = dt.strptime(day_str, fmt).date()
                    break
                except ValueError:
                    continue
            if parsed_date and parsed_date == today:
                today_row = row_idx
        except Exception:
            pass
        
        for col_idx, value in enumerate(row, start=1):
            cell = ws_schedule.cell(row=row_idx, column=col_idx, value=value)
            # Apply doctor color if this is a doctor cell (columns 2-10)
            if 2 <= col_idx <= 10 and value and value.strip() and value.strip() != "-":
                doc = value.strip()
                if doc in doctor_colors:
                    cell.fill = PatternFill(start_color=doctor_colors[doc], 
                                            end_color=doctor_colors[doc], 
                                            fill_type="solid")
            # Center align all cells
            cell.alignment = Alignment(horizontal="center")
    
    # Highlight today's row
    if today_row:
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws_schedule.cell(row=today_row, column=col_idx)
            cell.border = Border(top=Side(style='thick', color='FF0000'),
                                 bottom=Side(style='thick', color='FF0000'),
                                 left=Side(style='thick', color='FF0000'),
                                 right=Side(style='thick', color='FF0000'))
    
    # Add filter on first row
    ws_schedule.auto_filter.ref = ws_schedule.dimensions
    
    # Adjust column widths
    for col in ws_schedule.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_schedule.column_dimensions[column].width = adjusted_width
    
    # === Sheet 2: Counts & Fairness ===
    ws_counts = wb.create_sheet(title="Counts & Fairness")
    
    # Header
    headers = ["Doctor", "IM AM", "IM PM", "ICU AM", "ICU PM", "Hospitalist", "ECG", 
               "CV Clinic", "Dialysis", "Stress Tests", "Weekends", "Nights", "Friday Nights"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_counts.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, doc in enumerate(sorted(scheduler.all_doctors), start=2):
        stat = scheduler.stats[doc]
        ws_counts.cell(row=row_idx, column=1, value=doc)
        ws_counts.cell(row=row_idx, column=2, value=stat.im_am)
        ws_counts.cell(row=row_idx, column=3, value=stat.im_pm)
        ws_counts.cell(row=row_idx, column=4, value=stat.icu_am)
        ws_counts.cell(row=row_idx, column=5, value=stat.icu_pm)
        ws_counts.cell(row=row_idx, column=6, value=stat.hospitalist)
        ws_counts.cell(row=row_idx, column=7, value=stat.ecg)
        ws_counts.cell(row=row_idx, column=8, value=stat.cv_clinic)
        ws_counts.cell(row=row_idx, column=9, value=stat.dialysis)
        ws_counts.cell(row=row_idx, column=10, value=stat.stress_test)
        ws_counts.cell(row=row_idx, column=11, value=stat.weekends)
        ws_counts.cell(row=row_idx, column=12, value=stat.total_nights)
        ws_counts.cell(row=row_idx, column=13, value=stat.friday_nights)
    
    # Format Counts sheet
    for col in ws_counts.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws_counts.column_dimensions[column].width = adjusted_width
        # Center align numeric columns
        for cell in col:
            if cell.row > 1 and cell.column > 1:
                cell.alignment = Alignment(horizontal="center")
    
    # === Sheet 3: Doctor-Color Legend ===
    ws_legend = wb.create_sheet(title="Doctor Colors")
    ws_legend.cell(row=1, column=1, value="Doctor")
    ws_legend.cell(row=1, column=2, value="Color")
    ws_legend.cell(row=1, column=3, value="Hex Code")
    
    for col_idx in range(1, 4):
        cell = ws_legend.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, doc in enumerate(sorted(scheduler.all_doctors), start=2):
        ws_legend.cell(row=row_idx, column=1, value=doc)
        color_hex = doctor_colors[doc]
        ws_legend.cell(row=row_idx, column=2).fill = PatternFill(
            start_color=color_hex, end_color=color_hex, fill_type="solid")
        ws_legend.cell(row=row_idx, column=3, value=f"#{color_hex}")
    
    # Adjust legend column widths
    ws_legend.column_dimensions['A'].width = 25
    ws_legend.column_dimensions['B'].width = 15
    ws_legend.column_dimensions['C'].width = 12
    
    # Save workbook to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


# ============================================================================
# SIDEBAR
# ============================================================================

with st.sidebar:
    st.image("https://img.icons8.com/color/96/hospital-3.png", width=80)
    st.title("⚙️ Schedule Settings")
    
    st.subheader("📅 Schedule Period")
    
    col1, col2 = st.columns(2)
    with col1:
        year = st.selectbox("Year", [2025, 2026, 2027, 2028, 2029], index=2)
    
    with col2:
        # Default to Jan 1 of selected year
        default_start = datetime(year, 1, 1)
        start_date = st.date_input(
            "Start Date",
            value=default_start,
            help="First day of the schedule"
        )
    
    # Duration options
    duration_options = {
        "Full Year (365 days)": 365,
        "Full Year (366 days - leap year)": 366,
        "6 Months (182 days)": 182,
        "3 Months (91 days)": 91,
        "1 Month (30 days)": 30,
        "Custom": 0
    }
    
    duration_choice = st.selectbox(
        "Duration",
        list(duration_options.keys()),
        index=0,
        help="How many days to generate"
    )
    
    if duration_choice == "Custom":
        num_days = st.number_input(
            "Number of Days",
            min_value=7,
            max_value=730,
            value=365,
            step=7,
            help="Enter custom number of days (7-730)"
        )
    else:
        num_days = duration_options[duration_choice]
    
    # Show end date
    end_date = datetime.combine(start_date, datetime.min.time()) + timedelta(days=num_days - 1)
    st.caption(f"📆 End Date: **{end_date.strftime('%B %d, %Y')}**")
    
    # Auto-generate holidays when year changes
    current_holiday_year = None
    if st.session_state.holidays:
        # Get year from first holiday date
        first_date = list(st.session_state.holidays.keys())[0]
        try:
            current_holiday_year = int(first_date[:4])
        except:
            pass
    
    if current_holiday_year != year:
        st.session_state.holidays = generate_ns_holidays(year)
        st.toast(f"🗓️ Generated NS holidays for {year}")
    
    st.divider()
    
    st.subheader("📊 Historical Balancing")
    st.markdown("""
    Upload last year's fairness report to continue balancing assignments 
    across years. This helps ensure long-term fairness.
    """)
    
    uploaded_historical = st.file_uploader(
        "Upload Previous Fairness Report (CSV/TXT)",
        type=["csv", "txt"],
        help="Upload a fairness report from a previous schedule to continue balancing",
        key="historical_upload"
    )

    # Option B: Pull counts from an existing Google Sheet schedule
    with st.expander("Or load counts from Google Sheet"):
        st.markdown("""Paste a **publicly readable** Google Sheet link (Schedule grid).\n\nWe will compute historical assignment counts per doctor from the sheet, and use them as the starting point so the generator can **continue** the rotation fairly.""")
        default_sheet = "https://docs.google.com/spreadsheets/d/19JCZS_n9xOGik9uSByLmIvhI0P-MmfHLxCL69QX6WrA/edit?usp=sharing"
        sheet_url = st.text_input("Google Sheet URL", value=default_sheet, key="gsheet_url")
        gid = st.text_input("(Optional) Sheet tab gid", value="", key="gsheet_gid", help="Leave blank to use the default tab")

        colA, colB = st.columns(2)
        with colA:
            if st.button("🔄 Load Counts", key="load_counts_gsheet", use_container_width=True):
                try:
                    with st.spinner("Fetching Google Sheet and computing counts..."):
                        doctors_list = list(st.session_state.doctors.keys())
                        cutoff = datetime.combine(start_date, datetime.min.time())
                        counts = fetch_google_sheet_schedule_counts(sheet_url, gid.strip() or None, doctors_list, cutoff_date=cutoff)
                        st.session_state.historical_counts = counts
                    st.success(f"✓ Loaded counts from Google Sheet for {len(counts)} doctors")
                except Exception as e:
                    st.error(f"Failed to load from Google Sheet: {e}")

        with colB:
            st.caption("Tip: If names in the sheet differ (e.g., initials), update the Doctors list to match.")

    if uploaded_historical:
        # Parse the uploaded historical data
        historical_counts = parse_historical_report(uploaded_historical)
        if historical_counts:
            st.session_state.historical_counts = historical_counts
            st.success(f"✓ Loaded historical data for {len(historical_counts)} doctors")
        else:
            st.warning("Could not parse historical data. Make sure it's a fairness report.")

    use_historical = "historical_counts" in st.session_state and st.session_state.historical_counts
    
    if use_historical:
        with st.expander("View Historical Counts"):
            hist_df = pd.DataFrame(st.session_state.historical_counts).T
            st.dataframe(hist_df, use_container_width=True)
        
        if st.button("🗑️ Clear Historical Data"):
            del st.session_state.historical_counts
            st.rerun()
    
    st.divider()
    
    st.subheader("📤 Import Vacations")
    
    uploaded_vacations = st.file_uploader(
        "Upload Vacations (CSV/Excel)",
        type=["csv", "xlsx", "xls"],
        help="Columns: Doctor, StartDate, EndDate"
    )
    
    if uploaded_vacations:
        vacations = parse_vacation_csv(uploaded_vacations)
        if vacations:
            st.session_state.vacations = vacations
            st.success(f"✓ Loaded {len(vacations)} doctors' vacations")

    st.divider()
    if st.button("🚪 Logout"):
        st.session_state.authenticated = False
        st.session_state.username = None
        st.rerun()


# ============================================================================
# MAIN CONTENT
# ============================================================================

st.title("🏥 YRH Internal Medicine Schedule Generator")
st.markdown("Generate fair, conflict-free physician schedules with ease.")

# Create tabs
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["👨‍⚕️ Doctors & Services", "🔄 Rotation Rules", "🏖️ Vacations", "📅 Holidays", "📋 Generate Schedule", "ℹ️ Scheduling Rules"])

# ============================================================================
# TAB 1: DOCTORS & SERVICES
# ============================================================================

with tab1:
    st.subheader("Physician Service Coverage")
    st.markdown("Select which services each doctor can cover. Doctors with both **ICU** and **IM** can cover both services at night (reducing burnout).")
    
    # Convert to dataframe for editing
    df = doctors_to_dataframe(st.session_state.doctors)
    
    # Create columns for the table header
    col_widths = [2] + [1] * len(ALL_SERVICES) + [0.5]
    cols = st.columns(col_widths)
    cols[0].markdown("**Doctor**")
    for i, svc in enumerate(ALL_SERVICES):
        cols[i + 1].markdown(f"**{svc}**")
    cols[-1].markdown("**Del**")
    
    # Editable rows
    doctors_to_delete = []
    updated_doctors = {}
    
    for idx, row in df.iterrows():
        cols = st.columns(col_widths)
        
        # Doctor name (editable)
        new_name = cols[0].text_input(
            "Name", 
            value=row["Doctor"], 
            key=f"doc_name_{idx}",
            label_visibility="collapsed"
        )
        
        # Service checkboxes
        services = []
        for i, svc in enumerate(ALL_SERVICES):
            if cols[i + 1].checkbox(
                svc, 
                value=row.get(svc, False), 
                key=f"doc_{idx}_{svc}",
                label_visibility="collapsed"
            ):
                services.append(svc)
        
        # Delete button
        if cols[-1].button("🗑️", key=f"del_{idx}"):
            doctors_to_delete.append(row["Doctor"])
        
        if new_name and new_name not in doctors_to_delete:
            updated_doctors[new_name] = services
    
    # Add new doctor
    st.divider()
    col1, col2 = st.columns([3, 1])
    new_doc_name = col1.text_input("Add New Doctor", placeholder="Enter doctor name...")
    if col2.button("➕ Add Doctor", use_container_width=True):
        if new_doc_name and new_doc_name not in updated_doctors:
            updated_doctors[new_doc_name] = ["IM", "Hospitalist", "ECG"]  # Default services
            st.rerun()
    
    # Update session state
    st.session_state.doctors = updated_doctors
    
    # Show summary
    st.divider()
    icu_docs = [d for d, s in updated_doctors.items() if "ICU" in s]
    im_docs = [d for d, s in updated_doctors.items() if "IM" in s]
    dual_docs = [d for d in icu_docs if d in im_docs]
    
    col1, col2, col3 = st.columns(3)
    col1.metric("Total Doctors", len(updated_doctors))
    col2.metric("ICU Capable", len(icu_docs))
    col3.metric("Dual Coverage (ICU+IM)", len(dual_docs))


# ============================================================================
# TAB 2: ROTATION RULES
# ============================================================================

with tab2:
    st.subheader("Rotation Rules & Week Configuration")
    
    # Week Start Days
    st.markdown("### 📅 Week Start Days")
    st.markdown("Configure which day of the week each rotation starts on.")
    
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    col1, col2 = st.columns(2)
    
    with col1:
        im_icu_week_start = st.selectbox(
            "IM/ICU Week Starts On",
            day_names,
            index=st.session_state.week_starts["IM_ICU"],
            help="Day when IM and ICU weekly rotations begin"
        )
        st.session_state.week_starts["IM_ICU"] = day_names.index(im_icu_week_start)
    
    with col2:
        dialysis_week_start = st.selectbox(
            "Dialysis Week Starts On",
            day_names,
            index=st.session_state.week_starts["Dialysis"],
            help="Day when Dialysis weekly rotation begins"
        )
        st.session_state.week_starts["Dialysis"] = day_names.index(dialysis_week_start)
    
    st.divider()
    
    # Special Rotation Rules (Blair-type)
    st.markdown("### 🔁 Special Rotation Rules")
    st.markdown("""
    Define doctors with special ICU rotation patterns (like the "Blair Rule").
    These doctors only work ICU during specific weeks based on an interval.
    """)
    
    # Display current rules
    if st.session_state.blair_rules:
        st.markdown("**Current Special Rotations:**")
        
        rules_to_delete = []
        updated_rules = []
        
        for idx, rule in enumerate(st.session_state.blair_rules):
            with st.container():
                col1, col2, col3, col4, col5 = st.columns([3, 2, 3, 1, 1])
                
                with col1:
                    # Doctor selection
                    doctor_list = list(st.session_state.doctors.keys())
                    current_idx = doctor_list.index(rule["doctor"]) if rule["doctor"] in doctor_list else 0
                    new_doctor = st.selectbox(
                        "Doctor",
                        doctor_list,
                        index=current_idx,
                        key=f"blair_doc_{idx}",
                        label_visibility="collapsed" if idx > 0 else "visible"
                    )
                
                with col2:
                    # Interval (every N weeks)
                    new_interval = st.number_input(
                        "Every N Weeks",
                        min_value=2,
                        max_value=12,
                        value=rule.get("interval", 5),
                        key=f"blair_interval_{idx}",
                        label_visibility="collapsed" if idx > 0 else "visible"
                    )
                
                with col3:
                    # First Friday
                    try:
                        default_date = datetime.strptime(rule["first_friday"], "%Y-%m-%d").date()
                    except:
                        default_date = datetime(year, 1, 8).date()
                    
                    new_first_friday = st.date_input(
                        "First ICU Friday",
                        value=default_date,
                        key=f"blair_friday_{idx}",
                        label_visibility="collapsed" if idx > 0 else "visible"
                    )
                
                with col4:
                    st.write("")  # Spacer
                    if idx == 0:
                        st.write("")  # Extra spacer for label alignment
                    st.caption(f"1 in {new_interval}")
                
                with col5:
                    st.write("")  # Spacer
                    if idx == 0:
                        st.write("")  # Extra spacer for label alignment
                    if st.button("🗑️", key=f"del_blair_{idx}"):
                        rules_to_delete.append(idx)
                
                if idx not in rules_to_delete:
                    updated_rules.append({
                        "doctor": new_doctor,
                        "interval": new_interval,
                        "first_friday": new_first_friday.strftime("%Y-%m-%d")
                    })
        
        st.session_state.blair_rules = updated_rules
    
    # Add new rule
    st.divider()
    st.markdown("**Add New Special Rotation:**")
    
    col1, col2, col3, col4 = st.columns([3, 2, 3, 1])
    
    with col1:
        new_rule_doctor = st.selectbox(
            "Doctor",
            list(st.session_state.doctors.keys()),
            key="new_blair_doctor"
        )
    
    with col2:
        new_rule_interval = st.number_input(
            "Every N Weeks",
            min_value=2,
            max_value=12,
            value=5,
            key="new_blair_interval"
        )
    
    with col3:
        new_rule_friday = st.date_input(
            "First ICU Friday",
            value=datetime(year, 1, 8),
            key="new_blair_friday"
        )
    
    with col4:
        st.write("")  # Spacer for alignment
        if st.button("➕ Add", key="add_blair_rule", use_container_width=True):
            # Check if doctor already has a rule
            existing_doctors = [r["doctor"] for r in st.session_state.blair_rules]
            if new_rule_doctor in existing_doctors:
                st.warning(f"{new_rule_doctor} already has a special rotation rule.")
            else:
                st.session_state.blair_rules.append({
                    "doctor": new_rule_doctor,
                    "interval": new_rule_interval,
                    "first_friday": new_rule_friday.strftime("%Y-%m-%d")
                })
                st.rerun()
    
    # Summary
    st.divider()
    if st.session_state.blair_rules:
        st.markdown("**Summary:**")
        for rule in st.session_state.blair_rules:
            st.markdown(f"- **{rule['doctor']}**: ICU 1 week every **{rule['interval']} weeks**, starting {rule['first_friday']}")


# ============================================================================
# TAB 3: VACATIONS
# ============================================================================

with tab3:
    st.subheader("Physician Vacations")
    
    # Show current vacations
    if st.session_state.vacations:
        st.markdown("**Current Vacation Schedule:**")
        
        vac_data = []
        for doc, dates in st.session_state.vacations.items():
            if dates:
                sorted_dates = sorted(dates)
                # Group consecutive dates
                start = sorted_dates[0]
                end = sorted_dates[0]
                for d in sorted_dates[1:]:
                    if (d - end).days == 1:
                        end = d
                    else:
                        vac_data.append({
                            "Doctor": doc,
                            "Start": start.strftime("%Y-%m-%d"),
                            "End": end.strftime("%Y-%m-%d"),
                            "Days": (end - start).days + 1
                        })
                        start = d
                        end = d
                vac_data.append({
                    "Doctor": doc,
                    "Start": start.strftime("%Y-%m-%d"),
                    "End": end.strftime("%Y-%m-%d"),
                    "Days": (end - start).days + 1
                })
        
        vac_df = pd.DataFrame(vac_data)
        st.dataframe(vac_df, use_container_width=True, hide_index=True)
        
        if st.button("🗑️ Clear All Vacations"):
            st.session_state.vacations = {}
            st.rerun()
    else:
        st.info("No vacations loaded. Upload a CSV/Excel file in the sidebar or add manually below.")
    
    # Manual vacation entry
    st.divider()
    st.markdown("**Add Vacation Manually:**")
    
    col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
    
    vac_doctor = col1.selectbox(
        "Doctor",
        list(st.session_state.doctors.keys()),
        key="vac_doc_select"
    )
    vac_start = col2.date_input("Start Date", key="vac_start")
    vac_end = col3.date_input("End Date", key="vac_end")
    
    if col4.button("➕ Add", key="add_vac", use_container_width=True):
        if vac_doctor and vac_start and vac_end:
            if vac_doctor not in st.session_state.vacations:
                st.session_state.vacations[vac_doctor] = []
            
            current = datetime.combine(vac_start, datetime.min.time())
            end = datetime.combine(vac_end, datetime.min.time())
            while current <= end:
                if current not in st.session_state.vacations[vac_doctor]:
                    st.session_state.vacations[vac_doctor].append(current)
                current += timedelta(days=1)
            
            st.success(f"Added vacation for {vac_doctor}")
            st.rerun()


# ============================================================================
# TAB 4: HOLIDAYS
# ============================================================================

with tab4:
    st.subheader(f"Statutory Holidays - {year}")
    
    st.markdown("""
    Nova Scotia statutory holidays are **auto-generated** based on the selected year.
    You can add, remove, or modify holidays below.
    """)
    
    # Show current holidays
    holiday_data = [
        {"Date": date, "Holiday": name}
        for date, name in sorted(st.session_state.holidays.items())
    ]
    
    st.dataframe(pd.DataFrame(holiday_data), use_container_width=True, hide_index=True)
    
    # Add custom holiday
    st.divider()
    st.markdown("**Add Custom Holiday:**")
    col1, col2, col3 = st.columns([2, 2, 1])
    
    new_holiday_date = col1.date_input("Date", key="new_holiday_date")
    new_holiday_name = col2.text_input("Holiday Name", key="new_holiday_name")
    
    if col3.button("➕ Add Holiday", use_container_width=True):
        if new_holiday_date and new_holiday_name:
            date_str = new_holiday_date.strftime("%Y-%m-%d")
            st.session_state.holidays[date_str] = new_holiday_name
            st.rerun()
    
    # Remove holiday
    st.divider()
    st.markdown("**Remove Holiday:**")
    col1, col2 = st.columns([3, 1])
    
    holiday_to_remove = col1.selectbox(
        "Select holiday to remove",
        [f"{date} - {name}" for date, name in sorted(st.session_state.holidays.items())],
        key="holiday_to_remove"
    )
    
    if col2.button("🗑️ Remove", use_container_width=True):
        if holiday_to_remove:
            date_to_remove = holiday_to_remove.split(" - ")[0]
            if date_to_remove in st.session_state.holidays:
                del st.session_state.holidays[date_to_remove]
                st.rerun()
    
    # Regenerate for any year
    st.divider()
    col1, col2 = st.columns([2, 1])
    regen_year = col1.number_input("Generate NS Holidays for Year", min_value=2020, max_value=2050, value=year, key="regen_year")
    if col2.button("🔄 Generate", use_container_width=True):
        st.session_state.holidays = generate_ns_holidays(regen_year)
        st.success(f"Generated {len(st.session_state.holidays)} holidays for {regen_year}")
        st.rerun()


# ============================================================================
# TAB 5: GENERATE SCHEDULE
# ============================================================================

with tab5:
    st.subheader("Generate Schedule")
    
    # Summary
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Start Date", start_date.strftime("%b %d, %Y"))
    col2.metric("Duration", f"{num_days} days")
    col3.metric("Doctors", len(st.session_state.doctors))
    col4.metric("Vacations", sum(len(v) for v in st.session_state.vacations.values()))
    col5.metric("Holidays", len(st.session_state.holidays))
    
    st.divider()
    
    # Generate button
    if st.button("🚀 Generate Schedule", type="primary", use_container_width=True):
        with st.spinner("Generating 3 schedule options..."):
            
            # Convert holidays to datetime
            holidays_dt = {}
            for date_str, name in st.session_state.holidays.items():
                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                    holidays_dt[dt] = name
                except:
                    pass
            
            # Build config
            config = {
                "year": year,
                "start_date": datetime.combine(start_date, datetime.min.time()),
                "num_days": num_days,
                "doctors": st.session_state.doctors,
                "vacations": st.session_state.vacations,
                "holidays": holidays_dt,
                "blair_rules": st.session_state.blair_rules,
                "im_icu_week_start": st.session_state.week_starts["IM_ICU"],
                "dialysis_week_start": st.session_state.week_starts["Dialysis"],
                "use_historical": use_historical,
                "historical_counts": st.session_state.get("historical_counts", {}),
            }
            
            # Generate 3 options
            schedules = []
            for i in range(1, 4):
                random.seed(42 + i * 100 + year)
                scheduler = IMScheduler(config)
                scheduler.generate_schedule()
                schedules.append({
                    "option": i,
                    "scheduler": scheduler,
                    "df": scheduler.to_dataframe(),
                    "fairness": scheduler.get_fairness_report(),
                    "conflicts": scheduler.generate_conflict_report(),
                })
            
            st.session_state.generated_schedules = schedules
        
        st.success("✅ Generated 3 schedule options!")
    
    # Display results
    if st.session_state.generated_schedules:
        st.divider()
        
        for sched in st.session_state.generated_schedules:
            with st.expander(f"📋 Option {sched['option']}", expanded=(sched['option'] == 1)):
                
                # Conflict summary
                conflicts = sched['scheduler'].check_conflicts()
                num_conflicts = len([c for c in conflicts if c.startswith("CONFLICT")])
                num_warnings = len([c for c in conflicts if c.startswith("WARNING")])
                
                col1, col2, col3 = st.columns(3)
                col1.metric("Conflicts", num_conflicts, delta=None if num_conflicts == 0 else "Need attention", delta_color="inverse")
                col2.metric("Warnings", num_warnings)
                col3.metric("Days Scheduled", len(sched['df']))
                
                # Preview
                st.markdown("**Schedule Preview (First 2 Weeks):**")
                st.dataframe(sched['df'].head(14), use_container_width=True, hide_index=True)
                
                # Reports in tabs
                report_tab1, report_tab2 = st.tabs(["📊 Fairness Report", "⚠️ Conflict Report"])
                
                with report_tab1:
                    st.code(sched['fairness'], language=None)
                
                with report_tab2:
                    st.code(sched['conflicts'], language=None)
                
                # Download buttons
                st.markdown("**Download:**")
                col1, col2, col3, col4 = st.columns(4)
                
                # Schedule CSV
                csv_buffer = StringIO()
                sched['df'].to_csv(csv_buffer, index=False)
                col1.download_button(
                    "📥 Schedule (CSV)",
                    csv_buffer.getvalue(),
                    f"IM_Schedule_{year}_Option{sched['option']}.csv",
                    "text/csv",
                    use_container_width=True
                )
                
                # Schedule Excel (Formatted)
                excel_buffer = export_schedule_to_excel(sched['scheduler'])
                start_str = sched['scheduler'].start_date.strftime("%Y-%m-%d")
                end_date = sched['scheduler'].start_date + timedelta(days=sched['scheduler'].num_days - 1)
                end_str = end_date.strftime("%Y-%m-%d")
                col2.download_button(
                    "📊 Schedule (Excel)",
                    excel_buffer.getvalue(),
                    f"IM_Schedule_{start_str}_to_{end_str}_Option{sched['option']}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                # Fairness report
                col3.download_button(
                    "📥 Fairness Report",
                    sched['fairness'],
                    f"Fairness_Report_{year}_Option{sched['option']}.txt",
                    "text/plain",
                    use_container_width=True
                )
                
                # Conflict report
                col4.download_button(
                    "📥 Conflict Report",
                    sched['conflicts'],
                    f"Conflict_Report_{year}_Option{sched['option']}.txt",
                    "text/plain",
                    use_container_width=True
                )
        
        # Download all as ZIP
        st.divider()
        
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for sched in st.session_state.generated_schedules:
                # Add CSV
                csv_buffer = StringIO()
                sched['df'].to_csv(csv_buffer, index=False)
                zf.writestr(f"IM_Schedule_{year}_Option{sched['option']}.csv", csv_buffer.getvalue())
                
                # Add Excel (formatted)
                excel_buffer = export_schedule_to_excel(sched['scheduler'])
                start_str = sched['scheduler'].start_date.strftime("%Y-%m-%d")
                end_date = sched['scheduler'].start_date + timedelta(days=sched['scheduler'].num_days - 1)
                end_str = end_date.strftime("%Y-%m-%d")
                zf.writestr(f"IM_Schedule_{start_str}_to_{end_str}_Option{sched['option']}.xlsx", excel_buffer.getvalue())
                
                # Add reports
                zf.writestr(f"Fairness_Report_{year}_Option{sched['option']}.txt", sched['fairness'])
                zf.writestr(f"Conflict_Report_{year}_Option{sched['option']}.txt", sched['conflicts'])
        
        st.download_button(
            "📦 Download All (ZIP)",
            zip_buffer.getvalue(),
            f"IM_Schedules_{year}_All_Options.zip",
            "application/zip",
            use_container_width=True,
            type="primary"
        )


# ============================================================================
# TAB 6: SCHEDULING RULES INFO
# ============================================================================

with tab6:
    st.subheader("ℹ️ Scheduling Rules & Algorithm")
    
    st.markdown("""
    This page documents all the rules and logic used by the scheduling algorithm.
    Understanding these rules will help you configure the scheduler correctly.
    """)
    
    # Weekly Rotations
    st.markdown("### 📅 Weekly Rotation Rules")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **IM/ICU Weekly Rotation**
        - Rotates on configured day (default: Friday)
        - Same doctor covers IM or ICU for the entire week
        - Week runs from start day to day before next start day
        - Doctor on IM also covers ECG for that week
        
        **Hospitalist Assignment**
        - If ICU doctor can do Hospitalist → they cover it
        - Otherwise, a separate Hospitalist is assigned
        - Avoids double-booking IM and ICU doctors
        """)
    
    with col2:
        st.markdown("""
        **Dialysis Weekly Rotation**
        - Rotates on configured day (default: Monday)
        - Independent of IM/ICU rotation
        - Only doctors with Dialysis service assigned
        
        **Special Rotation Rules (Blair-type)**
        - Doctors work ICU only every N weeks
        - Not included in regular ICU rotation
        - Takes priority when their week comes up
        """)
    
    st.divider()
    
    # Night Call Rules
    st.markdown("### 🌙 Night Call Rules")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **Primary Night Service by Day**
        | Day | Primary Service |
        |-----|----------------|
        | Monday | IM |
        | Tuesday | ICU |
        | Wednesday | IM |
        | Thursday | ICU |
        | Friday | ICU |
        | Saturday | IM |
        | Sunday | ICU |
        """)
    
    with col2:
        st.markdown("""
        **Night Assignment Priority**
        1. If Blair-rule doctor's week → they take ICU nights
        2. Day doctor continues if able (reduces handoffs)
        3. Dual-coverage doctors preferred (cover both ICU+IM)
        4. Fairness-based selection among remaining candidates
        
        **Consecutive Night Limit**
        - Maximum 3 consecutive nights per doctor
        - Enforced by conflict checker
        - Generates WARNING if exceeded
        """)
    
    st.divider()
    
    # CV Clinic & Stress Tests
    st.markdown("### 🏥 CV Clinic & Stress Tests")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **CV Clinic**
        - Fridays only (except holidays)
        - Cannot be same doctor as IM or ICU that day
        - Selected from CVClinic-qualified doctors
        - Fairness-balanced across the year
        """)
    
    with col2:
        st.markdown("""
        **Stress Tests**
        - Monday through Thursday only
        - Shows "CV Clinic" on Fridays
        - Shows "WEEKEND" on Sat/Sun
        - Shows "HOLIDAY" on stat holidays
        - Cannot be same doctor as IM/ICU that day
        """)
    
    st.divider()
    
    # Fairness Algorithm
    st.markdown("### ⚖️ Fairness Algorithm")
    
    st.markdown("""
    The scheduler uses a **weighted scoring system** to ensure fair distribution:
    
    **Primary Score**: Count of assignments for the specific service being assigned
    
    **Penalty Weights** (added to primary score):
    - Weekend assignments: +0.5 per weekend day worked
    - Night assignments: +0.3 per night worked  
    - Friday nights: +0.4 per Friday night worked
    
    **Selection Process**:
    1. Calculate total score for each candidate
    2. Find candidates within 2 points of lowest score
    3. Randomly select from this "similar" group
    4. This adds variety while maintaining fairness
    
    **Historical Balancing**:
    - Upload previous year's fairness report
    - Starting counts are pre-loaded from historical data
    - Helps balance assignments across multiple years
    """)
    
    st.divider()
    
    # Conflict Detection
    st.markdown("### ⚠️ Conflict Detection")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **CONFLICTS (Must Fix)**
        - Doctor scheduled while on vacation
        - Missing coverage for required service
        - No ICU AM/PM coverage
        - No IM AM/PM coverage
        - No Hospitalist coverage
        - No Dialysis coverage
        """)
    
    with col2:
        st.markdown("""
        **WARNINGS (Review Recommended)**
        - More than 3 consecutive nights
        - More than 14 consecutive AM days
        - Missing CV Clinic on Friday
        - Missing Stress Test Mon-Thu
        
        *Warnings don't prevent schedule use but should be reviewed*
        """)
    
    st.divider()
    
    # Service Coverage
    st.markdown("### 👨‍⚕️ Service Coverage Categories")
    
    st.markdown("""
    Doctors are categorized based on their service assignments:
    
    | Category | Description | Night Call Behavior |
    |----------|-------------|---------------------|
    | **Dual Coverage** | Can do both ICU and IM | Preferred for nights (covers both services) |
    | **ICU Only** | Only ICU service | Only takes ICU nights |
    | **IM Only** | Only IM service | Only takes IM nights |
    | **ICU + Hospitalist** | ICU doctors who also do Hospitalist | Auto-assigned Hospitalist when on ICU |
    
    **Current Configuration:**
    """)
    
    # Show current doctor categories
    icu_docs = [d for d, s in st.session_state.doctors.items() if "ICU" in s]
    im_docs = [d for d, s in st.session_state.doctors.items() if "IM" in s]
    dual_docs = [d for d in icu_docs if d in im_docs]
    icu_only = [d for d in icu_docs if d not in im_docs]
    im_only = [d for d in im_docs if d not in icu_docs]
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"**Dual Coverage ({len(dual_docs)}):**")
        for d in dual_docs:
            st.markdown(f"- {d}")
    with col2:
        st.markdown(f"**ICU Only ({len(icu_only)}):**")
        for d in icu_only:
            st.markdown(f"- {d}")
    with col3:
        st.markdown(f"**IM Only ({len(im_only)}):**")
        for d in im_only:
            st.markdown(f"- {d}")
    
    st.divider()
    
    # Holiday Calculation
    st.markdown("### 📅 Nova Scotia Holiday Calculation")
    
    st.markdown("""
    Holidays are **automatically calculated** for any year:
    
    | Holiday | Rule |
    |---------|------|
    | New Year's Day | January 1 |
    | Heritage Day | 3rd Monday of February |
    | Good Friday | 2 days before Easter Sunday |
    | Easter Monday | 1 day after Easter Sunday |
    | Victoria Day | Monday before May 25 |
    | Canada Day | July 1 |
    | Civic Holiday | 1st Monday of August |
    | Labour Day | 1st Monday of September |
    | Truth & Reconciliation | September 30 |
    | Thanksgiving | 2nd Monday of October |
    | Remembrance Day | November 11 |
    | Christmas Day | December 25 |
    | Boxing Day | December 26 |
    
    *Easter is calculated using the Anonymous Gregorian algorithm*
    """)
    
    st.divider()
    
    # Tips
    st.markdown("### 💡 Tips for Best Results")
    
    st.markdown("""
    1. **Enter all vacations first** - This prevents scheduling conflicts
    2. **Review all 3 options** - Each uses different random seeds for variety
    3. **Check the Conflict Report** - Fix any CONFLICTS before using a schedule
    4. **Use Historical Balancing** - Upload last year's report for multi-year fairness
    5. **Adjust Blair rules carefully** - Overlapping intervals can cause issues
    6. **Regenerate if needed** - Run multiple times if no option is satisfactory
    """)


# ============================================================================
# FOOTER
# ============================================================================

st.divider()
st.markdown(
    """
    <div style='text-align: center; color: #888;'>
        YRH Internal Medicine Schedule Generator v2.0<br>
        Built for Yarmouth Regional Hospital Department of Internal Medicine
    </div>
    """,
    unsafe_allow_html=True
)
